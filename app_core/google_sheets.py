import re
from io import BytesIO

import openpyxl
import pandas as pd

from app_core.data import normalize_frame


SHEETS_READONLY_SCOPE = "https://www.googleapis.com/auth/spreadsheets.readonly"
DRIVE_READONLY_SCOPE = "https://www.googleapis.com/auth/drive.readonly"


def service_account_credentials(credentials_config, scopes):
    from google.oauth2.service_account import Credentials

    if isinstance(credentials_config, dict):
        return Credentials.from_service_account_info(dict(credentials_config), scopes=scopes)
    return Credentials.from_service_account_file(credentials_config, scopes=scopes)


def extract_spreadsheet_id(value):
    text = str(value or "").strip()
    match = re.search(r"/spreadsheets/d/([^/]+)", text)
    if match:
        return match.group(1)
    return text


def first_non_empty_header(values):
    header_tokens = {"week start", "week ending", "venue", "dollar sales", "unit sales"}
    for index, row in enumerate(values):
        normalized = {str(cell).strip().lower() for cell in row}
        if len(header_tokens & normalized) >= 3:
            return index
    for index, row in enumerate(values):
        if any(str(cell).strip() for cell in row):
            return index
    return 0


DEMO_FILL_COLORS = {"FF60CBF3"}


def is_demo_fill(cell):
    fill = cell.fill
    if not fill or not fill.fill_type:
        return False
    color = fill.fgColor
    return color.type == "rgb" and color.rgb in DEMO_FILL_COLORS


def unique_headers(headers):
    counts = {}
    output = []
    for header in headers:
        count = counts.get(header, 0)
        output.append(header if count == 0 else f"{header}.{count}")
        counts[header] = count + 1
    return output


def values_to_frame(values, demo_flags=None):
    if not values:
        return pd.DataFrame()
    header_index = first_non_empty_header(values)
    headers = values[header_index]
    rows = values[header_index + 1 :]
    width = max([len(headers), *(len(row) for row in rows)] or [0])
    headers = [str(headers[i]).strip() if i < len(headers) and str(headers[i]).strip() else f"Column {i + 1}" for i in range(width)]
    headers = unique_headers(headers)
    padded = [row + [None] * (width - len(row)) for row in rows]
    frame = pd.DataFrame(padded, columns=headers)
    if demo_flags is not None:
        frame["isDemoWeek"] = demo_flags[header_index + 1 : header_index + 1 + len(rows)]
    return frame


def worksheet_values_and_demo_flags(worksheet):
    values = []
    demo_flags = []
    for row in worksheet.iter_rows():
        values.append([cell.value for cell in row])
        demo_flags.append(any(is_demo_fill(cell) for cell in row))
    return values, demo_flags


def load_public_google_sheet(spreadsheet_id):
    url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"
    excel = pd.ExcelFile(url)
    frames = []
    errors = []
    for index, sheet_name in enumerate(excel.sheet_names):
        try:
            raw = pd.read_excel(excel, sheet_name=sheet_name)
            frame = normalize_frame(raw, f"Google Sheet: {spreadsheet_id}", sheet_name, source_order=index)
            frame["commonName"] = frame["commonName"].fillna(sheet_name)
            frame["productTab"] = sheet_name
            frames.append(frame)
        except Exception as exc:
            errors.append(f"{sheet_name}: {exc}")
    combined = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    combined.attrs["load_errors"] = errors
    return combined


def excel_bytes_to_frame(content, source_name):
    workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)
    frames = []
    errors = []
    warnings = []
    for index, sheet_name in enumerate(workbook.sheetnames):
        try:
            raw_values, demo_flags = worksheet_values_and_demo_flags(workbook[sheet_name])
            raw = values_to_frame(raw_values, demo_flags)
            frame = normalize_frame(raw, source_name, sheet_name, source_order=index)
            warnings.extend(frame.attrs.get("load_warnings", []))
            frame["commonName"] = frame["commonName"].fillna(sheet_name)
            frame["productTab"] = sheet_name
            if not frame.empty:
                frames.append(frame)
            elif sheet_name.strip().upper() != "SUMMARY":
                warnings.append(f"{sheet_name}: no usable data rows loaded.")
        except Exception as exc:
            errors.append(f"{sheet_name}: {exc}")
    combined = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    combined.attrs["load_errors"] = errors
    combined.attrs["load_warnings"] = warnings
    return combined


def load_drive_excel_file(file_id, credentials_path):
    from googleapiclient.discovery import build
    from googleapiclient.errors import HttpError
    from googleapiclient.http import MediaIoBaseDownload

    credentials = service_account_credentials(credentials_path, scopes=[DRIVE_READONLY_SCOPE])
    drive = build("drive", "v3", credentials=credentials, cache_discovery=False)
    try:
        metadata = drive.files().get(fileId=file_id, fields="id,name,mimeType").execute()
        request = drive.files().get_media(fileId=file_id)
        buffer = BytesIO()
        downloader = MediaIoBaseDownload(buffer, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
    except HttpError as exc:
        message = str(exc)
        if "SERVICE_DISABLED" in message or "drive.googleapis.com" in message and "disabled" in message.lower():
            raise RuntimeError(
                "Google Drive API is disabled for the service-account project. Enable it here: "
                "https://console.developers.google.com/apis/api/drive.googleapis.com/overview?project=731495116495"
            ) from exc
        raise RuntimeError(
            "Could not download the Office file through Google Drive. Confirm the Drive API is enabled and the "
            "file is shared with the service-account email."
        ) from exc

    return excel_bytes_to_frame(buffer.getvalue(), metadata.get("name", f"Google Drive file: {file_id}"))


def load_private_google_sheet(spreadsheet_id, credentials_path):
    try:
        from googleapiclient.discovery import build
        from googleapiclient.errors import HttpError
    except ImportError as exc:
        raise ImportError(
            "Google Sheets private access requires google-api-python-client and google-auth. "
            "Run setup_and_run.sh to install requirements."
        ) from exc

    credentials = service_account_credentials(credentials_path, scopes=[SHEETS_READONLY_SCOPE])
    service = build("sheets", "v4", credentials=credentials, cache_discovery=False)
    try:
        metadata = service.spreadsheets().get(spreadsheetId=spreadsheet_id, includeGridData=False).execute()
    except HttpError as exc:
        message = str(exc)
        if "must not be an Office file" in message:
            return load_drive_excel_file(spreadsheet_id, credentials_path)
        if "SERVICE_DISABLED" in message or "sheets.googleapis.com" in message and "disabled" in message.lower():
            raise RuntimeError(
                "Google Sheets API is disabled for the service-account project. Enable it here: "
                "https://console.developers.google.com/apis/api/sheets.googleapis.com/overview?project=731495116495"
            ) from exc
        if exc.resp.status in {401, 403, 404}:
            raise RuntimeError(
                "The service account cannot access this Sheet. Share the Sheet with the service-account email "
                "from the JSON key, then retry."
            ) from exc
        raise
    sheet_titles = [sheet["properties"]["title"] for sheet in metadata.get("sheets", [])]

    frames = []
    errors = []
    warnings = []
    for index, title in enumerate(sheet_titles):
        try:
            result = service.spreadsheets().values().get(spreadsheetId=spreadsheet_id, range=f"'{title}'").execute()
            raw = values_to_frame(result.get("values", []))
            frame = normalize_frame(raw, f"Google Sheet: {spreadsheet_id}", title, source_order=index)
            warnings.extend(frame.attrs.get("load_warnings", []))
            frame["commonName"] = frame["commonName"].fillna(title)
            frame["productTab"] = title
            frames.append(frame)
            if frame.empty and title.strip().upper() != "SUMMARY":
                warnings.append(f"{title}: no usable data rows loaded.")
        except Exception as exc:
            errors.append(f"{title}: {exc}")

    combined = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    combined.attrs["load_errors"] = errors
    combined.attrs["load_warnings"] = warnings
    return combined


def load_google_sheet(spreadsheet_url_or_id, credentials_path=None):
    spreadsheet_id = extract_spreadsheet_id(spreadsheet_url_or_id)
    if credentials_path:
        return load_private_google_sheet(spreadsheet_id, credentials_path)
    return load_public_google_sheet(spreadsheet_id)
